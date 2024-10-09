import dimensions
import metrics
import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# Configurazione della pagina con titolo e icona
st.set_page_config(page_title="Keyword Cannibalization Tool", page_icon="🔑")

# Aggiungi il logo di Nur Digital Marketing
st.image("./logo_nur_vettoriale.svg", width=100)  # Sostituisci l'URL con il percorso del logo

# Titolo e descrizione dello strumento
st.markdown("# Keyword Cannibalization Tool")
st.markdown("""
Trova le pagine che competono tra loro per la stessa parola chiave utilizzando i dati GSC.
I dati devono contenere informazioni multidimensionali su query (parole chiave) e pagine.
""")

# Aggiunta delle indicazioni in una sezione collapsabile (st.expander)
with st.expander("Come funziona?"):
    st.markdown("""
1.  Vai sullo strumento [GSC API](https://search-console-connector.streamlit.app/)
2.  Clicca su "Sign-in with Google"
3.  Dopo aver fatto l'autenticazione, clicca su "Access GSC API"
4.  Seleziona il dominio da "Web property to review (please sign in via Google OAuth first)"
5.  Dimension: lascia "query"
6.  Nested dimension: seleziona "page"
7.  Seleziona un periodo temporale dal menu a discesa
8.  Clicca su "Fetch GSC API Data"
9.  Scarica il file da "Download CSV"
10.  Torna sullo strumento di cannibalizzazione e carica il file appena scaricato da GSC
    """)   

# Slider per impostare la soglia
perc_slider = st.slider('Imposta soglia (es: 80 = Selezione delle prime 80% di query per metrica)', 0, 100, value=80, step=10, key='perc_slider')

# Spiegazione della soglia
st.markdown("""
<div style="background-color: #F0F2F6; padding: 10px; border-radius: 5px;">
    <span style="font-size: 12px;">
    <b>ℹ️ Cosa significa la soglia?</b><br>
    La soglia che imposti indica che l'analisi si concentrerà solo sulle query o parole chiave che generano la percentuale impostata del totale delle prestazioni. 
    Ad esempio, una soglia dell'80% significa che verranno analizzate solo le query che rappresentano l'80% del traffico o delle metriche selezionate, 
    ignorando quelle meno rilevanti.
    </span>
</div>
""", unsafe_allow_html=True)
    
st.markdown("---")

# Caricamento del file CSV
gsc_data_file = st.file_uploader('Carica Dati', type='csv', key='key')

summary = []

# Funzione per controllare la fonte dei dati
def check_source(df):
    if 'query' in df.columns.str.lower() and 'clicks' in df.columns.str.lower():
        if 'page' in df.columns.str.lower():
            return dimensions.gsc_dimensions, metrics.gsc_metrics
        else:
            st.error('I dati devono contenere una colonna "page".')
            st.stop()
    elif 'keyword' in df.columns.str.lower() and 'url' in df.columns.str.lower():
        return dimensions.semrush_dimensions, metrics.semrush_metrics
    else:
        st.error('File dati non valido. Si prega di caricare un file valido.')
        st.stop()

# Funzione per filtrare righe con caratteri non ASCII
def is_ascii(string):
    try:
        string.encode('ascii')
    except UnicodeEncodeError:
        return False
    else:
        return True

# Funzione per elaborare i dati e calcolare le metriche
def process_data(df, metric, perc_cumsum, dimension):
    dimension = [d.lower() for d in dimension]
    groupby_column_name = dimension[0]
    metric_name = metric.lower()

    if 'query' in dimension:
        df_cols = [groupby_column_name, 'page', metric_name, 'ctr', 'position', metric_name + '_percent_all_query',
                   'query_percentile_' + metric_name]
    if 'keyword' in dimension:
        df_cols = [groupby_column_name, 'url', metric_name, 'cpc', 'position', metric_name + '_percent_all_keyword']

    grouped_df = df.groupby(metric_name, as_index=False).apply(
        lambda group: group.sort_values(metric_name, ascending=False))

    sum_per_group = grouped_df.groupby(groupby_column_name)[metric_name].sum().sort_values(ascending=False)
    percent_per_group = grouped_df.groupby(groupby_column_name)[metric_name].sum() / grouped_df[metric_name].sum()

    cumsum_percent = percent_per_group.sort_values(ascending=False).cumsum()
    top_n_percent = percent_per_group[cumsum_percent <= float(perc_cumsum)]

    df = pd.merge(grouped_df, sum_per_group, left_on=groupby_column_name, right_index=True, suffixes=('', '_sum'))
    df = pd.merge(df, top_n_percent, left_on=groupby_column_name, right_index=True,
                  suffixes=('', f'_percent_all_{dimension[0]}'))

    df.sort_values([metric_name + '_sum', metric_name], ascending=[False, False], inplace=True)
    df[f'{dimension[0]}_percentile_' + metric_name] = df[metric_name] / df[metric_name + '_sum']
    df.sort_values([metric_name + '_sum', metric_name], ascending=[False, False], inplace=True)
    df = df[df[f'{dimension[0]}_percentile_' + metric_name] >= 0.10]
    df.drop_duplicates(subset=[dimension[0], dimension[1]], inplace=True)
    df = df[df.duplicated(subset=[dimension[0]], keep=False)]
    df.sort_values([metric_name + '_sum', dimension[0], metric_name, 'position'], ascending=[False, True, False, True],
                   inplace=True)

    unique_values = df[dimension[0]].nunique()
    summary.append(f'{unique_values} parole chiave uniche in competizione per {metric_name}.')
    df = df[df_cols]
    return df

# Funzione per unire i DataFrame elaborati
def process_merge(dfs, dimension):
    merged_df = pd.merge(dfs[0], dfs[1], on=[dimension[0], dimension[1]], how='inner')
    merged_df = merged_df[merged_df.duplicated(subset=[dimension[0]], keep=False)]
    merged_df = merged_df.loc[:, ~merged_df.columns.str.endswith('_y')]
    merged_df = merged_df.rename(columns=lambda x: x.replace('_x', ''))
    return merged_df

# Funzione per formattare l'Excel, aggiungendo le righe in verde
def format_excel(xlsx_file):
    wb = load_workbook(filename=xlsx_file)
    
    # Formattazione per le righe "buone"
    good_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_font = Font(color="006100")

    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = 130
        seen_queries = set()

        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=False):
            query = row[0].value
            if query not in seen_queries:
                seen_queries.add(query)
                for cell in row:
                    cell.fill = good_fill
                    cell.font = good_font

        ws.title = f'Competing by {ws.title}'

    wb.save(f'cannibalization_data_threshold_{perc_slider}.xlsx')
    return wb

# Main
if __name__ == '__main__':
    if gsc_data_file is not None:
        data = pd.read_csv(gsc_data_file)
        data.columns = data.columns.str.lower()

        dimensions, metrics = check_source(data)
        data = data[data[dimensions[0]].apply(lambda x: is_ascii(str(x)))]
        perc_cumsum = perc_slider / 100

        data = data[data[metrics[0]] > 0]
        dfs = []
        wb = Workbook()

        for metric in metrics[:2]:
            df_processed = process_data(data, metric, perc_cumsum, dimensions)
            dfs.append(df_processed)

        dfs.append(process_merge(dfs, dimensions))
        unique_vals = dfs[2][dimensions[0]].nunique()
        summary.append(f'{unique_vals} parole chiave uniche in competizione per {metrics[2]}.')

        try:
            for sheet_name, df in zip(metrics, dfs):
                sheet = wb.create_sheet(title=sheet_name)
                for row in dataframe_to_rows(df, index=False, header=True):
                    sheet.append(row)
        except IndexError:
            st.error('Soglia troppo bassa. Si prega di aumentare la soglia.')
        
        wb.remove(wb['Sheet'])
        wb.save(f'cannibalization_data_threshold_{perc_slider}.xlsx')

        wb = format_excel(f'cannibalization_data_threshold_{perc_slider}.xlsx')
        wb.save(f'cannibalization_data_threshold_{perc_slider}.xlsx')

        st.markdown(f'### Riassunto, soglia impostata al {perc_slider}%')
        for s in summary:
            st.markdown(f'{s}')
        
        with open(f"cannibalization_data_threshold_{perc_slider}.xlsx", "rb") as file:
            st.download_button(label='Scarica report di cannibalizzazione',
                               data=file,
                               file_name=f'cannibalization_data_threshold_{perc_slider}.xlsx',
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               key='download_button')

# Copyright and footer
st.markdown("---")
st.markdown("© 2024 [NUR® Digital Marketing](https://www.nur.it/). Tutti i diritti riservati.")
